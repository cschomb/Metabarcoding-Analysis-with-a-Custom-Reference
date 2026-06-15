#!/usr/bin/env python3
"""
Metabarcoding pipeline: BLAST → Analysis → GBIF occurrence check.

Workflow:
    1. BLAST    – builds a local nucleotide DB from a reference FASTA and runs
                  blastn on ESV (Exact Sequence Variant) query sequences.
    2. Analysis – filters / scores BLAST hits, picks the best hit per ESV,
                  and joins results back to the ESV abundance table.
    3. GBIF     – resolves taxonomy against the GBIF backbone, counts occurrences
                  inside a user-defined geographic circle, and annotates each ESV
                  as "plausible" / "implausible" based on local occurrence data.
                  Occurrences are always queried via the GBIF API.
                  Providing credentials (gbif_user / gbif_pwd) enables higher
                  concurrency and more lenient rate limits.

Usage:
    python3 pipeline.py init <project_name>
    python3 pipeline.py run  --project <project_name> --steps all
"""
from __future__ import annotations

import random
import argparse
import json
import math
import os
import re
import subprocess
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font
from pyproj import Geod
from requests.adapters import HTTPAdapter
from tqdm import tqdm
from urllib3.util.retry import Retry

# ---------------------------------------------------------------------------
# Global constants
# ---------------------------------------------------------------------------

RETRYABLE_STATUS = {429, 500, 502, 503, 504}
UNKNOWN_TAXON    = "NA"
RUN_TS           = datetime.now().strftime("%Y%m%d_%H%M%S")

# ---------------------------------------------------------------------------
# Logging helpers
# ---------------------------------------------------------------------------

def log_tqdm(paths: ProjectPaths, msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    with open(paths.logfile, "a", encoding="utf-8") as fh:
        fh.write(line + "\n")
    tqdm.write(line)


# ---------------------------------------------------------------------------
# Project-scoped path management
# ---------------------------------------------------------------------------

class ProjectPaths:
    """
    Central registry of every file-system location used by the pipeline.

        <root>/
            input/
            01_blast/tmp/, 01_blast/results/
            02_analysis/results/
            03_gbif/results/
            settings.xlsx
            log.txt
    """
    def __init__(self, root: Path):
        self.root             = root
        self.input            = self.root / "input"
        self.blast_dir        = self.root / "01_blast"
        self.blast_tmp        = self.blast_dir / "tmp"
        self.blast_results    = self.blast_dir / "results"
        self.analysis_dir     = self.root / "02_analysis"
        self.analysis_results = self.analysis_dir / "results"
        self.gbif_dir         = self.root / "03_gbif"
        self.gbif_results     = self.gbif_dir / "results"
        self.settings         = self.root / "settings.xlsx"
        self.logfile          = self.root / "log.txt"


def resolve_path_setting(paths: ProjectPaths, value: Any, default_dir: Path) -> Path:
    if _is_empty(value):
        raise ValueError("Expected a path value but got empty")
    s = str(value).strip()
    p = Path(s)
    if p.is_absolute():
        return p
    if ("/" in s) or ("\\" in s):
        return (paths.root / p).resolve()
    return (default_dir / p).resolve()


def resolve_path_setting_optional(
    paths: ProjectPaths, value: Any, default_dir: Path
) -> Optional[Path]:
    if _is_empty(value):
        return None
    return resolve_path_setting(paths, value, default_dir)


def log(paths: ProjectPaths, msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    paths.root.mkdir(parents=True, exist_ok=True)
    with open(paths.logfile, "a", encoding="utf-8") as fh:
        fh.write(line + "\n")


def console(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tqdm.write(f"[{ts}] {msg}")


def ensure_structure(paths: ProjectPaths) -> None:
    for p in [
        paths.input,
        paths.blast_tmp, paths.blast_results,
        paths.analysis_results,
        paths.gbif_results,
    ]:
        p.mkdir(parents=True, exist_ok=True)
    if not paths.logfile.exists():
        paths.logfile.write_text("", encoding="utf-8")


def log_settings(paths: ProjectPaths, title: str, settings: dict[str, Any]) -> None:
    log(paths, f"--- {title} settings ---")
    for k in sorted(settings.keys()):
        v = settings[k]
        if k == "gbif_pwd" and v and not _is_empty(v):
            v = "***"
        if isinstance(v, float) and np.isnan(v):
            v = ""
        log(paths, f"{k}: {v}")


# ---------------------------------------------------------------------------
# settings.xlsx template writer and reader
# ---------------------------------------------------------------------------

def write_settings_template(paths: ProjectPaths, force: bool = False) -> None:
    """
    Create a fresh settings.xlsx with three sheets (blast / analysis / gbif).

    GBIF occurrences are always counted via the API (occurrence/search endpoint).

    Without credentials (gbif_user / gbif_pwd left blank):
        - workers_unauth controls concurrency (default 4 → 16 threads)
        - Anonymous rate limits apply; keep rps at 4.0 or lower.

    With credentials (gbif_user + gbif_pwd provided):
        - workers_auth controls concurrency (default 8 → 32 threads)
        - Authenticated rate limits are more lenient; rps can be raised to 8+.
        - Credentials can be set in settings.xlsx or via environment variables
          GBIF_USER and GBIF_PWD.
    """
    if paths.settings.exists() and not force:
        raise FileExistsError(
            f"{paths.settings} already exists. Use --force to overwrite."
        )

    blast = pd.DataFrame(
        [
            {"parameter": "reference_fasta", "value": "reference.fasta",
             "description": "FASTA in input/ used to build the BLAST database (makeblastdb -dbtype nucl)."},
            {"parameter": "esv_fasta",        "value": "esv.fasta",
             "description": "Query FASTA in input/. FASTA containing the ESV sequences."},
            {"parameter": "max_target_seqs",  "value": 50,
             "description": "Maximum number of target sequences from the BLAST analysis."},
        ],
        columns=["parameter", "value", "description"],
    )

    analysis = pd.DataFrame(
        [
            {"parameter": "esv_table",       "value": "esv_table.xlsx",
             "description": "Excel ESV table in input/. The output from Apscale"},
            {"parameter": "score_file",       "value": "scoring.csv",
             "description": "Optional scoring table in input/. Leave empty for none."},
            {"parameter": "score_type",       "value": "",
             "description": "If score_file is used: 'all', for the complete database, or a column name from the scoring file (for Germany 'Score_ger')."},
            {"parameter": "length_threshold", "value": 200,
             "description": "Minimum alignment length to keep while analysing BLAST hits."},
            {"parameter": "blast_results",    "value": "",
             "description": "Optional override path to BLAST output. If empty, uses 01_blast/results/blast.out."},
        ],
        columns=["parameter", "value", "description"],
    )

    gbif = pd.DataFrame(
        [
            # --- General ---
            {"parameter": "input_results_xlsx", "value": "",
             "description": "Optional override. If blank, uses results from step 2."},
            {"parameter": "year_from",    "value": "",
             "description": "Count occurrences from this year onwards (inclusive)."},
            {"parameter": "lon",          "value": "",
             "description": "Longitude of circle center (WGS84)."},
            {"parameter": "lat",          "value": "",
             "description": "Latitude of circle center (WGS84)."},
            {"parameter": "radius_m",     "value": "",
             "description": "Radius in meters for the GBIF search circle."},
            {"parameter": "min_occurrences", "value": 5,
             "description": "Min occurrences required for occurs_in_area=True."},
            {"parameter": "n_points",     "value": 60,
             "description": "Polygon points to approximate the search circle (default 60)."},

            # --- Concurrency ---
            {"parameter": "workers_unauth", "value": 4,
             "description": (
                 "Worker threads when running WITHOUT credentials.\n"
                 "Actual thread count = workers_unauth * 4, capped at 16.\n"
                 "Applies to both taxonomy matching and occurrence counting.\n"
                 "Keep at 4 to stay within anonymous rate limits."
             )},
            {"parameter": "workers_auth",   "value": 8,
             "description": (
                 "Worker threads when running WITH credentials.\n"
                 "Actual thread count = workers_auth * 4, capped at 32.\n"
                 "Applies to both taxonomy matching and occurrence counting.\n"
                 "Can be raised to 16 on fast connections."
             )},

            # --- Rate limit ---
            {"parameter": "rps",          "value": 4.0,
             "description": (
                 "Soft target requests/second (used for logging).\n"
                 "With credentials you can safely raise this to 8.0 or higher.\n"
                 "Without credentials keep at 4.0 or lower."
             )},

            # --- Credentials (optional — enables higher concurrency) ---
            {"parameter": "gbif_user",    "value": "",
             "description": (
                 "GBIF username (free account at gbif.org).\n"
                 "Optional — providing credentials enables higher concurrency\n"
                 "and more lenient API rate limits.\n"
                 "Can also be set via GBIF_USER environment variable."
             )},
            {"parameter": "gbif_pwd",     "value": "",
             "description": (
                 "GBIF password.\n"
                 "Optional — see gbif_user.\n"
                 "Can also be set via GBIF_PWD environment variable."
             )},
        ],
        columns=["parameter", "value", "description"],
    )

    with pd.ExcelWriter(paths.settings, engine="openpyxl") as xw:
        blast.to_excel(xw, sheet_name="blast", index=False)
        analysis.to_excel(xw, sheet_name="analysis", index=False)
        gbif.to_excel(xw, sheet_name="gbif", index=False)

    force_value_column_text_format(paths.settings, ["blast", "analysis", "gbif"])


def read_sheet_kv(paths: ProjectPaths, sheet: str) -> dict[str, Any]:
    df = pd.read_excel(paths.settings, sheet_name=sheet)
    if df.shape[1] < 2:
        raise ValueError(f"Sheet '{sheet}' must have at least two columns: parameter, value")
    df = df.iloc[:, :2].copy()
    df.columns = ["parameter", "value"]
    df = df.dropna(subset=["parameter"])
    return {str(k).strip(): v for k, v in zip(df["parameter"], df["value"])}


# ---------------------------------------------------------------------------
# Settings value helpers
# ---------------------------------------------------------------------------

def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, float) and np.isnan(v):
        return True
    s = str(v).strip()
    return s == "" or s.upper() in {"NA", "NONE", "N/A", "NULL", "NAN"}


def get_str(d: dict[str, Any], k: str, default: Optional[str] = None) -> str:
    if k not in d or _is_empty(d[k]):
        if default is None:
            raise ValueError(f"Missing required setting '{k}'")
        return default
    return str(d[k]).strip()


def get_int(d: dict[str, Any], k: str, default: Optional[int] = None) -> int:
    if k not in d or _is_empty(d[k]):
        if default is None:
            raise ValueError(f"Missing required setting '{k}'")
        return default
    return int(d[k])


def get_float(d: dict[str, Any], k: str, default: Optional[float] = None) -> float:
    if k not in d or _is_empty(d[k]):
        if default is None:
            raise ValueError(f"Missing required setting '{k}'")
        return default
    return float(d[k])


def force_value_column_text_format(xlsx_path: Path, sheets: list[str]) -> None:
    wb = load_workbook(xlsx_path)
    for sh in sheets:
        ws = wb[sh]
        headers = {
            str(c.value).strip().lower(): i + 1
            for i, c in enumerate(ws[1])
            if c.value is not None
        }
        if "value" not in headers:
            continue
        col = headers["value"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "@"
    wb.save(xlsx_path)


# ---------------------------------------------------------------------------
# Subprocess helper
# ---------------------------------------------------------------------------

def run_cmd(paths: ProjectPaths, cmd: list[str]) -> None:
    log(paths, "Running: " + " ".join(cmd))
    try:
        p = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            check=True,
        )
        if p.stdout:
            for line in p.stdout.splitlines():
                log(paths, line)
    except FileNotFoundError as e:
        raise RuntimeError(
            f"Command not found: {cmd[0]}. Install BLAST+ and ensure it is on PATH."
        ) from e
    except subprocess.CalledProcessError as e:
        if e.stdout:
            for line in e.stdout.splitlines():
                log(paths, line)
        raise RuntimeError(
            f"Command failed (exit={e.returncode}): {' '.join(cmd)}"
        ) from e


# ---------------------------------------------------------------------------
# FASTA utilities
# ---------------------------------------------------------------------------

def fasta_records(path: Path):
    header = None
    seq_lines: list[str] = []
    with open(path, "r", encoding="utf-8") as fh:
        for line in fh:
            line = line.rstrip("\n")
            if line.startswith(">"):
                if header is not None:
                    yield header, "".join(seq_lines)
                header = line
                seq_lines = []
            else:
                if header is None:
                    continue
                seq_lines.append(line.strip())
        if header is not None:
            yield header, "".join(seq_lines)


def write_fasta_chunk(records: list[tuple[str, str]], out_path: Path) -> None:
    with open(out_path, "w", encoding="utf-8") as fh:
        for h, s in records:
            fh.write(h + "\n")
            fh.write(s + "\n")


def split_fasta_into_chunks(
    paths: ProjectPaths, query_fasta: Path, chunk_size: int = 100
) -> list[Path]:
    for p in paths.blast_tmp.glob("chunk_*.fasta"):
        p.unlink(missing_ok=True)

    chunks: list[Path] = []
    buf: list[tuple[str, str]] = []
    idx = 0
    nseq = 0

    for rec in fasta_records(query_fasta):
        buf.append(rec)
        nseq += 1
        if len(buf) >= chunk_size:
            idx += 1
            out = paths.blast_tmp / f"chunk_{idx:06d}.fasta"
            write_fasta_chunk(buf, out)
            chunks.append(out)
            buf = []

    if buf:
        idx += 1
        out = paths.blast_tmp / f"chunk_{idx:06d}.fasta"
        write_fasta_chunk(buf, out)
        chunks.append(out)

    log(paths, f"Split query into {len(chunks)} chunks (chunk_size={chunk_size}, sequences={nseq}).")
    return chunks


# ---------------------------------------------------------------------------
# Step 1: Local BLAST
# ---------------------------------------------------------------------------

BLAST_OUTFMT = (
    "6 qseqid sseqid pident length mismatch gapopen "
    "qstart qend sstart send evalue bitscore qcovs"
)


def step1_blast(paths: ProjectPaths, blast_settings: dict[str, Any]) -> Path:
    log(paths, "STEP 1: RUNNING BLAST")

    ref_fasta       = resolve_path_setting(paths, blast_settings["reference_fasta"], paths.input)
    query_fasta     = resolve_path_setting(paths, blast_settings["esv_fasta"],       paths.input)
    max_target_seqs = get_int(blast_settings, "max_target_seqs")
    num_threads     = os.cpu_count() or 1
    db_prefix       = paths.blast_tmp / "DB"
    out_file        = paths.blast_results / f"blast_{RUN_TS}.out"

    if not ref_fasta.exists():
        raise FileNotFoundError(f"reference_fasta not found: {ref_fasta}")
    if not query_fasta.exists():
        raise FileNotFoundError(f"esv_fasta not found: {query_fasta}")

    run_cmd(paths, ["makeblastdb", "-in", str(ref_fasta), "-dbtype", "nucl", "-out", str(db_prefix)])

    chunks = split_fasta_into_chunks(paths, query_fasta, chunk_size=100)
    if out_file.exists():
        out_file.unlink()

    log(paths, f"Running blastn chunked (threads={num_threads}, max_target_seqs={max_target_seqs})")

    for chunk in tqdm(chunks, desc="BLAST chunks", unit="chunks"):
        cmd = [
            "blastn",
            "-query",           str(chunk),
            "-db",              str(db_prefix),
            "-outfmt",          BLAST_OUTFMT,
            "-max_target_seqs", str(max_target_seqs),
            "-num_threads",     str(num_threads),
        ]
        try:
            p = subprocess.run(
                cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, check=True
            )
        except subprocess.CalledProcessError as e:
            log(paths, f"blastn failed for {chunk}")
            for stream_name, stream_content in [
                ("blastn stdout", e.stdout), ("blastn stderr", e.stderr)
            ]:
                if stream_content:
                    log(paths, f"--- {stream_name} ---")
                    for line in stream_content.splitlines():
                        log(paths, line)
            raise

        with open(out_file, "a", encoding="utf-8") as out:
            if p.stdout:
                out.write(p.stdout)
        if p.stderr:
            for line in p.stderr.splitlines():
                log(paths, line)

    for p in paths.blast_tmp.glob("chunk_*.fasta"):
        p.unlink(missing_ok=True)

    log(paths, f"BLAST finished. Output: {out_file}")
    return out_file


def cleanup_blast_db(paths: ProjectPaths) -> None:
    for f in paths.blast_tmp.glob("DB.*"):
        f.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Step 2: Analysis
# ---------------------------------------------------------------------------

def extract_taxonomy(tax_str):
    if pd.isnull(tax_str):
        return [None] * 8
    tax_parts = str(tax_str).replace("tax=", "").split(",")
    return [
        tax_parts[i].split(":")[1]
        if len(tax_parts) > i and ":" in tax_parts[i]
        else None
        for i in range(8)
    ]


def filter_blast(paths: ProjectPaths, input_file: Path, length_threshold: int) -> Path:
    output_csv = paths.analysis_results / f"filtered_blast_results_{RUN_TS}.csv"
    if output_csv.exists():
        output_csv.unlink()

    column_order = [
        "qseqid", "s_acc",
        "s_phylum", "s_class", "s_order", "s_family", "s_genus", "s_species",
        "gbif_tax", "ncbi_tax",
        "pident", "length", "mismatches", "gapopens",
        "qstart", "qend", "sstart", "send",
        "evalue", "bitscore", "qcov", "pident_interval",
    ]

    write_header = True
    chunksize    = 10_000

    try:
        total_lines = sum(1 for _ in open(input_file, encoding="utf-8", errors="ignore"))
        n_chunks    = math.ceil(total_lines / chunksize)
    except OSError:
        n_chunks = None

    log(paths, "Filtering BLAST results.")

    def process_group(group_df: pd.DataFrame, q_acc_name: str) -> Optional[pd.DataFrame]:
        n_input    = len(group_df)
        group_df   = group_df[group_df["length"] >= length_threshold].copy()
        if group_df.empty:
            console(f"qseqid {q_acc_name}: {n_input} hits, kept 0.")
            return None

        group_df["pident"] = group_df["pident"].round(3)
        top_pident         = group_df["pident"].max()
        pident_cutoff      = top_pident - 3.0
        selected_group     = group_df[group_df["pident"] >= pident_cutoff].copy()

        if selected_group.empty:
            console(f"qseqid {q_acc_name}: {n_input} hits, kept 0 after pident selection.")
            return None

        selected_group = selected_group.sort_values("pident", ascending=False)
        selected_group["pident_interval"] = f"{top_pident:.3f}-{pident_cutoff:.3f}"
        console(
            f"qseqid {q_acc_name}: {n_input} hits, kept {len(selected_group)}, "
            f"top pident {top_pident:.3f}, cutoff {pident_cutoff:.3f}"
        )
        return selected_group

    curr_q_acc = None
    buffer: list = []

    for chunk in tqdm(
        pd.read_csv(input_file, sep="\t", header=None, chunksize=chunksize),
        total=n_chunks, desc="Filtering BLAST chunks", dynamic_ncols=True, leave=True,
    ):
        chunk.columns = [
            "qseqid", "sseqid", "pident", "length", "mismatches", "gapopens",
            "qstart", "qend", "sstart", "send", "evalue", "bitscore", "qcov",
        ]
        chunk[["s_acc", "s_tax"]] = (
            chunk["sseqid"].astype(str).str.split(";", n=1, expand=True)
        )
        chunk[
            ["s_phylum", "s_class", "s_order", "s_family",
             "s_genus", "s_species", "gbif_tax", "ncbi_tax"]
        ] = chunk["s_tax"].apply(extract_taxonomy).apply(pd.Series)
        chunk.drop(["sseqid", "s_tax"], axis=1, inplace=True)
        chunk["length"] = pd.to_numeric(chunk["length"], errors="coerce")

        for _, row in chunk.iterrows():
            this_q = row["qseqid"]
            if curr_q_acc is None:
                curr_q_acc = this_q
            if this_q != curr_q_acc:
                filtered = process_group(pd.DataFrame(buffer), str(curr_q_acc))
                if filtered is not None and not filtered.empty:
                    filtered = filtered[column_order]
                    filtered.to_csv(output_csv, mode="a", header=write_header, index=False)
                    write_header = False
                buffer = []
                curr_q_acc = this_q
            buffer.append(row)

    if buffer:
        filtered = process_group(pd.DataFrame(buffer), str(curr_q_acc))
        if filtered is not None and not filtered.empty:
            filtered = filtered[column_order]
            filtered.to_csv(output_csv, mode="a", header=write_header, index=False)

    log(paths, f"Filtered BLAST CSV: {output_csv}")
    return output_csv


def score_blast(
    paths: ProjectPaths,
    csv_file: Path,
    scoring_file: Optional[Path],
    score_type: str,
) -> pd.DataFrame:
    tqdm.write("Scoring BLAST results...")
    tqdm.pandas()

    filtered_df    = pd.read_csv(csv_file)
    filtered_blast = filtered_df.copy()
    filtered_df["comment"] = ""

    def score_row(row):
        pident   = row["pident"]
        qcov     = row["qcov"]
        score    = (pident / 2) + (qcov / 2)
        taxonomy = {
            "s_phylum":  row["s_phylum"],
            "s_class":   row["s_class"],
            "s_order":   row["s_order"],
            "s_family":  row["s_family"],
            "s_genus":   row["s_genus"],
            "s_species": row["s_species"],
            "gbif_tax":  row["gbif_tax"],
            "ncbi_tax":  row["ncbi_tax"],
        }
        if pident > 97.0:
            pass
        elif pident > 94.0:
            taxonomy["s_species"] = taxonomy["gbif_tax"] = taxonomy["ncbi_tax"] = UNKNOWN_TAXON
        elif pident > 90.0:
            taxonomy["s_species"] = taxonomy["s_genus"] = UNKNOWN_TAXON
            taxonomy["gbif_tax"]  = taxonomy["ncbi_tax"] = UNKNOWN_TAXON
        else:
            for k in taxonomy:
                taxonomy[k] = UNKNOWN_TAXON
        return pd.Series({"blast_score": score, **taxonomy})

    taxonomy_cols = [
        "s_phylum", "s_class", "s_order", "s_family",
        "s_genus", "s_species", "gbif_tax", "ncbi_tax",
    ]
    score_cols = ["blast_score"] + taxonomy_cols

    if scoring_file is None:
        filtered_df[score_cols] = filtered_df.progress_apply(score_row, axis=1)
        merged_df = filtered_df
        merged_df["total_score"] = merged_df["blast_score"].fillna(0)
    else:
        scoring_df = pd.read_csv(scoring_file, sep=";")
        filtered_df[score_cols] = filtered_df.progress_apply(score_row, axis=1)
        merged_df = pd.merge(
            filtered_df, scoring_df,
            left_on="s_acc", right_on="Accession_Number", how="left",
        )
        if score_type == "local":
            extra_col = "score_local"
        elif score_type == "all":
            extra_col = "Score"
        else:
            extra_col = score_type

        if extra_col not in merged_df.columns:
            raise ValueError(
                f"Requested scoring column '{extra_col}' not found in scoring file.\n"
                "Set score_type to 'local', 'all', or to an existing column name."
            )
        merged_df["total_score"] = (
            merged_df["blast_score"].fillna(0) + merged_df[extra_col].fillna(0)
        )

    results: list                    = []
    other_candidates_list: list[str] = []

    grouped = merged_df.groupby("qseqid")
    for qseqid, group in tqdm(
        grouped, total=merged_df["qseqid"].nunique(), desc="Selecting best hits"
    ):
        group     = group.sort_values("total_score", ascending=False)
        max_score = group["total_score"].max()
        top_hits  = group[group["total_score"] == max_score]

        comment = ""
        if len(top_hits) > 1 and top_hits["blast_score"].nunique(dropna=False) == 1:
            if any(top_hits[c].nunique(dropna=False) != 1 for c in taxonomy_cols):
                comment = "multiple candidates with same blast score"

        all_candidates = filtered_blast[filtered_blast["qseqid"] == qseqid].copy()
        best_rows_set  = set(zip(top_hits["s_species"], top_hits["pident"]))
        mask           = all_candidates.apply(
            lambda r: (r["s_species"], r["pident"]) not in best_rows_set, axis=1
        )
        other_candidates_df = all_candidates[mask]

        if not other_candidates_df.empty:
            candidate_map: dict[str, set] = {}
            for _, r in other_candidates_df.iterrows():
                sp = r["s_species"]
                if pd.notna(sp) and sp not in [UNKNOWN_TAXON, None]:
                    candidate_map.setdefault(sp, set()).add(r["pident"])
            formatted = [
                f"{sp} {{{', '.join(f'{p:.3f}' for p in sorted(pidents, reverse=True))}}}"
                for sp, pidents in sorted(candidate_map.items())
            ]
            other_candidates_str = "; ".join(formatted) if formatted else "None"
        else:
            other_candidates_str = "None"

        if len(top_hits) == 1 or all(
            top_hits[c].nunique(dropna=False) == 1 for c in taxonomy_cols
        ):
            row_out = top_hits.iloc[0].copy()
            row_out["comment"] = comment
            results.append(row_out)
        else:
            unified_row = top_hits.iloc[0].copy()
            for col in taxonomy_cols:
                vals = top_hits[col].unique()
                unified_row[col] = vals[0] if len(vals) == 1 else UNKNOWN_TAXON
            unified_row["comment"] = comment
            results.append(unified_row)

        other_candidates_list.append(other_candidates_str)

    final_df = pd.DataFrame(results)
    final_df["other_candidates"] = other_candidates_list

    keep_cols = [
        "qseqid", "s_acc",
        "s_phylum", "s_class", "s_order", "s_family", "s_genus", "s_species",
        "gbif_tax", "ncbi_tax",
        "pident", "length", "mismatches", "gapopens",
        "qstart", "qend", "sstart", "send",
        "evalue", "bitscore", "qcov",
        "total_score", "quality",
        "comment", "other_candidates",
    ]
    final_df = final_df.loc[:, [c for c in keep_cols if c in final_df.columns]]

    out_best = paths.analysis_results / f"best_blast_hits_{RUN_TS}.xlsx"
    log(paths, f"Writing best BLAST hits: {out_best}")
    final_df.to_excel(out_best, index=False)
    return final_df


def results_file(paths: ProjectPaths, esv_xlsx: Path, scored: pd.DataFrame) -> Path:
    ESVs = pd.read_excel(esv_xlsx).sort_values(by="hash")
    df   = pd.merge(scored, ESVs, left_on="qseqid", right_on="hash", how="outer")
    df   = df.drop("qseqid", axis=1)
    df   = df.loc[:, ~df.columns.duplicated()]

    columns_to_remove = [
        "length", "mismatches", "gapopens", "qstart", "qend",
        "sstart", "send", "bitscore", "qcov",
    ]
    df = df.drop(columns=[c for c in columns_to_remove if c in df.columns])

    taxonomy_cols = [
        "s_acc", "s_phylum", "s_class", "s_order",
        "s_family", "s_genus", "s_species", "gbif_tax", "ncbi_tax",
    ]
    reserved_cols = (
        ["hash", "sequence", "pident", "evalue", "total_score",
         "comment", "quality", "sum_of_reads", "other_candidates"]
        + taxonomy_cols
    )

    sample_cols = [c for c in ESVs.columns if c not in reserved_cols and c in df.columns]
    df["sum_of_reads"] = df[sample_cols].sum(axis=1)

    results_order = (
        ["hash", "sequence"]
        + taxonomy_cols
        + ["pident", "evalue", "total_score", "comment", "quality",
           "other_candidates", "sum_of_reads"]
        + sample_cols
    )
    df = df[[col for col in results_order if col in df.columns]]
    df = df.replace(r"^\s*$", np.nan, regex=True).fillna("NA").infer_objects(copy=False)

    for col in ["s_species", "other_candidates"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace("_", " ", regex=False)

    tax_sort_cols = [
        c for c in ["s_phylum", "s_class", "s_order", "s_family", "s_genus", "s_species"]
        if c in df.columns
    ]
    df = df.sort_values(
        by=tax_sort_cols, ascending=True, na_position="last", kind="stable"
    ).reset_index(drop=True)

    out = paths.analysis_results / f"results_{RUN_TS}.xlsx"
    log(paths, f"Writing results table: {out}")
    df.to_excel(out, index=False)
    return out


def step2_analysis(
    paths: ProjectPaths,
    analysis_settings: dict[str, Any],
    blast_out_default: Path,
) -> Path:
    log(paths, "STEP 2: ANALYSIS")

    esv_table = resolve_path_setting(paths, analysis_settings["esv_table"], paths.input)

    score_file = None
    if "score_file" in analysis_settings and not _is_empty(analysis_settings["score_file"]):
        score_file = resolve_path_setting_optional(
            paths, analysis_settings.get("score_file"), paths.input
        )

    score_type       = get_str(analysis_settings, "score_type", default="all")
    length_threshold = get_int(analysis_settings, "length_threshold", default=200)

    blast_override = analysis_settings.get("blast_results")
    blast_file = (
        blast_out_default
        if _is_empty(blast_override)
        else resolve_path_setting(paths, blast_override, paths.input)
    )

    if not esv_table.exists():
        raise FileNotFoundError(f"ESV table not found: {esv_table}")
    if score_file is not None and not score_file.exists():
        raise FileNotFoundError(f"Score file not found: {score_file}")
    if not blast_file.exists():
        raise FileNotFoundError(f"BLAST results not found: {blast_file}")

    filtered_csv = filter_blast(paths, blast_file, length_threshold=length_threshold)
    scored_df    = score_blast(paths, filtered_csv, scoring_file=score_file, score_type=score_type)
    results_xlsx = results_file(paths, esv_table, scored_df)

    log(paths, "Analysis finished.")
    return results_xlsx


# ---------------------------------------------------------------------------
# Step 3: GBIF occurrence check
# ---------------------------------------------------------------------------

GBIF_API = "https://api.gbif.org/v1/"
HEADERS  = {"User-Agent": "MetabarcodingGBIFOccCheck/2.1"}

GREEN = "008000"
RED   = "FF0000"

_BRACE_RE = re.compile(r"\{[^}]*\}")


def _is_transient_error(note: str) -> bool:
    if not note:
        return False
    transient_phrases = [
        "Request failed",
        "Failed after 5 attempts",
        "ConnectionError",
        "Timeout",
        "timed out",
    ]
    if any(phrase in note for phrase in transient_phrases):
        return True
    return any(f"HTTP {code}" in note for code in RETRYABLE_STATUS)


# ---------------------------------------------------------------------------
# GBIF session factory + thread-local session management
# ---------------------------------------------------------------------------

def _make_gbif_session(
    gbif_user: Optional[str] = None,
    gbif_pwd: Optional[str] = None,
) -> requests.Session:
    """
    Build a requests.Session for GBIF API calls.
    If credentials are provided the session uses HTTP Basic Auth,
    which enables higher rate limits on the GBIF API.
    """
    session = requests.Session()
    session.headers.update(HEADERS)
    session.headers["Connection"] = "close"
    if gbif_user and gbif_pwd:
        session.auth = (gbif_user, gbif_pwd)

    retry = Retry(
        total=5,
        backoff_factor=2.0,
        status_forcelist=tuple(RETRYABLE_STATUS),
        allowed_methods=frozenset(["GET"]),
        raise_on_status=False,
        respect_retry_after_header=True,
    )
    adapter = HTTPAdapter(
        max_retries=retry,
        pool_connections=1,
        pool_maxsize=1,
    )
    session.mount("https://", adapter)
    session.mount("http://",  adapter)
    return session


_thread_local = threading.local()


def _get_thread_session(
    gbif_user: Optional[str] = None,
    gbif_pwd: Optional[str] = None,
) -> requests.Session:
    """
    Return a per-thread requests.Session, creating one if needed.
    Credentials are baked into the session at creation time.
    """
    if not hasattr(_thread_local, "session"):
        _thread_local.session = _make_gbif_session(gbif_user, gbif_pwd)
    return _thread_local.session


# ---------------------------------------------------------------------------
# Text / name cleaning utilities
# ---------------------------------------------------------------------------

def _clean_text_cell(x) -> str | None:
    if x is None or pd.isna(x):
        return None
    s = str(x).strip()
    if s == "" or s.upper() in {"NA", "N/A", "NULL", "NONE", "NAN"}:
        return None
    return s


def _clean_species_name(s: str | None) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = _BRACE_RE.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _parse_other_candidates(cell) -> list[str]:
    if pd.isna(cell) or cell is None:
        return []
    parts = [p.strip() for p in str(cell).split(";")]
    return [name for p in parts if (name := _clean_species_name(p))]


# ---------------------------------------------------------------------------
# GBIF taxonomy extraction
# ---------------------------------------------------------------------------

IN_RANK_COLS = ["s_species", "s_genus", "s_family", "s_order", "s_class", "s_phylum"]
OUT_COLS     = ["scientificName", "rank", "phylum", "class", "order", "family", "genus"]


def extract_taxonomy_table_gbif(paths: ProjectPaths, xlsx_path: Path) -> pd.DataFrame:
    log(paths, f"Extracting taxonomy table from: {xlsx_path}")
    df = pd.read_excel(xlsx_path, sheet_name=0)

    required = IN_RANK_COLS + ["other_candidates"]
    missing  = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    base = df[IN_RANK_COLS].copy()
    for c in IN_RANK_COLS:
        base[c] = base[c].map(_clean_text_cell)

    base = base.rename(columns={
        "s_species": "species", "s_genus": "genus", "s_family": "family",
        "s_order": "order",     "s_class": "class", "s_phylum": "phylum",
    })
    for c in ["species", "genus", "family", "order", "class", "phylum"]:
        base[c] = base[c].map(_clean_species_name)

    def _add_taxon(rows, rank, name, ctx):
        if not name:
            return
        rows.append({
            "scientificName": name,
            "rank":    rank,
            "phylum":  ctx.get("phylum") or None,
            "class":   ctx.get("class")  or None,
            "order":   ctx.get("order")  or None,
            "family":  ctx.get("family") or None,
            "genus":   ctx.get("genus")  or None,
        })

    expanded: list[dict[str, Any]] = []
    for _, r in base.iterrows():
        ctx = {lvl: r.get(lvl) or None for lvl in ("phylum", "class", "order", "family", "genus")}
        _add_taxon(expanded, "species", r.get("species") or "", ctx)
        _add_taxon(expanded, "genus",   r.get("genus")   or "", ctx)
        _add_taxon(expanded, "family",  r.get("family")  or "", ctx)

    out = pd.DataFrame(expanded, columns=OUT_COLS)

    existing = set(out["scientificName"].dropna().tolist())
    extras: list[dict[str, Any]] = []
    for cell in df["other_candidates"]:
        for name in _parse_other_candidates(cell):
            if name in existing:
                continue
            existing.add(name)
            extras.append({
                "scientificName": name, "rank": "species",
                "phylum": None, "class": None, "order": None,
                "family": None, "genus": None,
            })
    if extras:
        out = pd.concat([out, pd.DataFrame(extras, columns=OUT_COLS)], ignore_index=True)

    out = (
        out.dropna(subset=["scientificName"])
           .drop_duplicates(subset=["scientificName"], keep="first")
           .reset_index(drop=True)
    )
    log(paths, f"Extracted {len(out)} unique taxa")
    return out


def is_base_row(row: pd.Series) -> bool:
    return any(
        pd.notna(row.get(col)) and str(row.get(col)).strip() != ""
        for col in ["phylum", "class", "order", "family", "genus"]
    )


def _taxonomy_key(taxonomy: Optional[dict[str, str | None]]) -> tuple:
    if not taxonomy:
        return ()
    return (
        taxonomy.get("phylum"),
        taxonomy.get("class"),
        taxonomy.get("order"),
        taxonomy.get("family"),
        taxonomy.get("genus"),
    )


_MATCH_CACHE: dict[tuple, dict[str, Any]] = {}
_CACHE_LOCK = threading.Lock()


# ---------------------------------------------------------------------------
# GBIF taxonomy matching
# ---------------------------------------------------------------------------

def match_taxon_to_gbif(
    name: str,
    rank: Optional[str] = None,
    taxonomy: Optional[dict] = None,
    timeout_s: int = 30,
    gbif_user: Optional[str] = None,
    gbif_pwd: Optional[str] = None,
) -> dict[str, Any]:
    """
    Match a single taxon name against the GBIF species/match endpoint.
    Uses the thread-local session (with credentials if provided).
    """
    cache_key = (name, rank, _taxonomy_key(taxonomy))
    with _CACHE_LOCK:
        if cache_key in _MATCH_CACHE:
            return _MATCH_CACHE[cache_key]

    params: dict[str, str] = {"name": name, "verbose": "false"}
    if rank:
        params["rank"] = str(rank).upper()
    if taxonomy:
        for k in ("phylum", "class", "order", "family"):
            if taxonomy.get(k):
                params[k] = str(taxonomy[k])
        if taxonomy.get("genus") and str(rank or "").lower() != "genus":
            params["genus"] = str(taxonomy["genus"])

    url      = f"{GBIF_API}species/match"
    last_exc = None

    for attempt in range(5):
        try:
            r = _get_thread_session(gbif_user, gbif_pwd).get(
                url, params=params, timeout=timeout_s
            )

            if r.status_code != 200:
                if r.status_code not in RETRYABLE_STATUS:
                    out = {
                        "gbif_taxon_key": None, "gbif_match_type": None,
                        "gbif_scientific_name": None, "gbif_rank": None,
                        "confidence": None,
                        "note": f"HTTP {r.status_code}: {r.text[:200]}",
                        "alternatives": [], "accepted_key": None, "accepted_name": None,
                    }
                    with _CACHE_LOCK:
                        _MATCH_CACHE[cache_key] = out
                    return out
                sleep_s = (2 ** attempt) + random.uniform(0, 1)
                tqdm.write(
                    f"  match '{name}' attempt {attempt + 1} "
                    f"HTTP {r.status_code} — retrying in {sleep_s:.1f}s"
                )
                time.sleep(sleep_s)
                continue

            data = r.json()

            if data.get("matchType") == "HIGHERRANK" and not data.get("usageKey"):
                if data.get("acceptedUsageKey"):
                    data["usageKey"]  = data["acceptedUsageKey"]
                    data["matchType"] = "HIGHERRANK_WITH_KEY"
                else:
                    data["usageKey"]  = None
                    data["matchType"] = "HIGHERRANK_NO_KEY"

            result = {
                "gbif_taxon_key":       data.get("usageKey"),
                "gbif_match_type":      data.get("matchType"),
                "gbif_scientific_name": data.get("scientificName"),
                "gbif_rank":            data.get("rank"),
                "confidence":           data.get("confidence"),
                "note":                 data.get("note"),
                "alternatives":         data.get("alternatives", []),
                "accepted_key":         data.get("acceptedUsageKey"),
                "accepted_name":        data.get("acceptedUsage"),
            }
            with _CACHE_LOCK:
                _MATCH_CACHE[cache_key] = result
            return result

        except requests.exceptions.RequestException as e:
            last_exc = e
            sleep_s = (2 ** attempt) + random.uniform(0, 1)
            tqdm.write(
                f"  match '{name}' attempt {attempt + 1} failed: {e} "
                f"— retrying in {sleep_s:.1f}s"
            )
            time.sleep(sleep_s)

    out = {
        "gbif_taxon_key": None, "gbif_match_type": None,
        "gbif_scientific_name": None, "gbif_rank": None,
        "confidence": None,
        "note": f"Request failed after 5 attempts: {last_exc}",
        "alternatives": [], "accepted_key": None, "accepted_name": None,
    }
    return out


def match_taxonomy_with_gbif(
    paths: ProjectPaths,
    taxonomy_df: pd.DataFrame,
    safe_workers: int,
    gbif_user: Optional[str] = None,
    gbif_pwd: Optional[str] = None,
) -> pd.DataFrame:
    """
    Match all taxa against the GBIF backbone using a thread pool.
    safe_workers is already computed by the caller based on credentials.
    """
    log(paths, "Matching taxonomy against GBIF backbone")
    log(paths, f"Using {safe_workers} worker threads for taxonomy matching")

    rows_by_name = taxonomy_df.set_index("scientificName", drop=False).to_dict(orient="index")

    taxa_to_process = []
    for _, row in taxonomy_df.iterrows():
        taxonomy = None
        if is_base_row(row):
            taxonomy = {k: row.get(k) for k in ("phylum", "class", "order", "family", "genus")}
        taxa_to_process.append((row["scientificName"], row.get("rank"), taxonomy))

    def process_one(args):
        name, rank, tax = args
        return name, rank, tax, match_taxon_to_gbif(
            name, rank, tax, gbif_user=gbif_user, gbif_pwd=gbif_pwd
        )

    results: list[dict[str, Any]] = []

    with ThreadPoolExecutor(max_workers=safe_workers) as ex:
        futures = [ex.submit(process_one, t) for t in taxa_to_process]
        for fut in tqdm(
            as_completed(futures),
            total=len(futures),
            desc="Matching taxa",
            unit="taxon",
        ):
            name, rank, tax, match = fut.result()
            o = rows_by_name[name]
            results.append({
                "scientificName":    name,
                "original_rank":     o.get("rank"),
                "original_phylum":   o.get("phylum"),
                "original_class":    o.get("class"),
                "original_order":    o.get("order"),
                "original_family":   o.get("family"),
                "original_genus":    o.get("genus"),
                **match,
                "alternatives_json": json.dumps(
                    match.get("alternatives", []), ensure_ascii=False
                ),
            })

    failed = [
        (r["scientificName"], r.get("original_rank"),
         {k.replace("original_", ""): r.get(k)
          for k in ("original_phylum", "original_class", "original_order",
                    "original_family", "original_genus")})
        for r in results
        if r.get("gbif_taxon_key") is None
        and _is_transient_error(str(r.get("note") or ""))
    ]

    if failed:
        log(paths, f"Retrying {len(failed)} transient taxonomy failures after cooldown")
        time.sleep(15)
        for name, rank, tax in tqdm(failed, desc="Retry taxonomy", unit="taxon"):
            match = match_taxon_to_gbif(
                name, rank, tax, gbif_user=gbif_user, gbif_pwd=gbif_pwd
            )
            for i, row in enumerate(results):
                if row["scientificName"] == name and row.get("original_rank") == rank:
                    results[i] = {
                        **results[i], **match,
                        "alternatives_json": json.dumps(
                            match.get("alternatives", []), ensure_ascii=False
                        ),
                    }
                    break

    return pd.DataFrame(results)


# ---------------------------------------------------------------------------
# GBIF occurrence — shared helpers
# ---------------------------------------------------------------------------

def build_wkt_circle(lon: float, lat: float, radius_m: float, n_points: int = 60) -> str:
    geod     = Geod(ellps="WGS84")
    bearings = np.linspace(0, 360, n_points, endpoint=True)
    pts      = []
    for b in bearings:
        x, y, _ = geod.fwd(lon, lat, b, radius_m)
        pts.append((x, y))
    pts = list(reversed(pts))
    if pts and pts[0] != pts[-1]:
        pts.append(pts[0])
    coords = ", ".join(f"{x} {y}" for x, y in pts)
    return f"POLYGON(({coords}))"


def _write_checkpoint(
    out_df: pd.DataFrame,
    key_to_result: dict[int, dict[str, Any]],
    checkpoint_path: Path,
) -> None:
    tmp = out_df.copy()
    for i, row in tmp.iterrows():
        k = row.get("gbif_taxon_key")
        if pd.isna(k):
            tmp.at[i, "occ_error"] = "No gbif_taxon_key"
            continue
        kk = int(k)
        if kk in key_to_result:
            res = key_to_result[kk]
            tmp.at[i, "occurs_in_area"]   = res.get("occurs_in_area")
            tmp.at[i, "occurrence_count"] = res.get("occurrence_count")
            tmp.at[i, "occ_error"]        = res.get("occ_error")
    tmp.to_excel(checkpoint_path, index=False)


def _apply_occurrence_results_vectorised(
    out: pd.DataFrame,
    key_to_result: dict[int, dict[str, Any]],
) -> pd.DataFrame:
    no_key_mask = out["gbif_taxon_key"].isna()
    out.loc[no_key_mask, "occ_error"] = "No gbif_taxon_key"

    valid = out.loc[~no_key_mask].copy()
    valid["_int_key"] = valid["gbif_taxon_key"].astype(int)

    valid["occurs_in_area"] = valid["_int_key"].map(
        lambda k: key_to_result.get(k, {}).get("occurs_in_area")
    )
    valid["occurrence_count"] = valid["_int_key"].map(
        lambda k: key_to_result.get(k, {}).get("occurrence_count")
    )
    valid["occ_error"] = valid["_int_key"].map(
        lambda k: (
            key_to_result.get(k, {}).get("occ_error")
            or ("No result for gbif_taxon_key" if k not in key_to_result else None)
        )
    )
    valid = valid.drop(columns=["_int_key"])
    out.loc[~no_key_mask] = valid
    return out


# ---------------------------------------------------------------------------
# GBIF occurrence counting via API
# ---------------------------------------------------------------------------

def count_occurrences_api(
    paths: ProjectPaths,
    matches_df: pd.DataFrame,
    wkt_geometry: str,
    year_from: int,
    min_occurrences: int,
    safe_workers: int,
    checkpoint_path: Optional[Path],
    gbif_user: Optional[str] = None,
    gbif_pwd: Optional[str] = None,
) -> pd.DataFrame:
    """
    Count GBIF occurrences using the occurrence/search endpoint with a
    thread pool. safe_workers is already computed by the caller.
    Each thread reuses its own session (with credentials if provided).
    """
    out = matches_df.copy()
    out["occurs_in_area"]   = None
    out["occurrence_count"] = None
    out["occ_error"]        = None

    unique_keys = sorted(int(k) for k in out["gbif_taxon_key"].dropna().unique())
    if not unique_keys:
        log(paths, "No taxon keys to query.")
        return out

    log(
        paths,
        f"Counting occurrences via API: {len(unique_keys)} keys, "
        f"{safe_workers} workers"
    )

    url              = f"{GBIF_API}occurrence/search"
    key_to_result:   dict[int, dict[str, Any]] = {}
    result_lock      = threading.Lock()
    checkpoint_every = 200

    def fetch_one(taxon_key: int) -> tuple[int, dict[str, Any]]:
        params = {
            "taxonKey":           str(taxon_key),
            "geometry":           wkt_geometry,
            "year":               f"{int(year_from)},*",
            "occurrenceStatus":   "PRESENT",
            "hasGeospatialIssue": "false",
            "hasCoordinate":      "true",
            "limit":              "0",
            "offset":             "0",
        }
        last_exc = None
        for attempt in range(5):
            try:
                r = _get_thread_session(gbif_user, gbif_pwd).get(
                    url, params=params, timeout=30
                )

                if r.status_code == 200:
                    data   = r.json()
                    cnt    = data.get("count", None)
                    occurs = (cnt >= int(min_occurrences)) if isinstance(cnt, int) else None
                    return taxon_key, {
                        "occurrence_count": cnt,
                        "occurs_in_area":   occurs,
                        "occ_error":        None,
                    }
                elif r.status_code in RETRYABLE_STATUS:
                    retry_after = r.headers.get("Retry-After")
                    sleep_s = (
                        float(retry_after) if retry_after
                        else (2 ** attempt) + random.uniform(1, 3)
                    )
                    tqdm.write(
                        f"  key {taxon_key} attempt {attempt + 1} "
                        f"HTTP {r.status_code} — sleeping {sleep_s:.1f}s"
                    )
                    time.sleep(sleep_s)
                else:
                    return taxon_key, {
                        "occurrence_count": None,
                        "occurs_in_area":   None,
                        "occ_error":        f"HTTP {r.status_code}: {r.text[:200]}",
                    }
            except requests.exceptions.RequestException as e:
                last_exc = e
                sleep_s  = (2 ** attempt) + random.uniform(0, 2)
                tqdm.write(
                    f"  key {taxon_key} attempt {attempt + 1} "
                    f"failed: {e} — retrying in {sleep_s:.1f}s"
                )
                time.sleep(sleep_s)

        return taxon_key, {
            "occurrence_count": None,
            "occurs_in_area":   None,
            "occ_error":        f"Failed after 5 attempts: {last_exc}",
        }

    completed = [0]

    with ThreadPoolExecutor(max_workers=safe_workers) as ex:
        futures = {ex.submit(fetch_one, k): k for k in unique_keys}
        for fut in tqdm(
            as_completed(futures),
            total=len(futures),
            desc="Occurrence counts",
            unit="taxon",
        ):
            k, res = fut.result()
            with result_lock:
                key_to_result[k] = res
                completed[0] += 1
                if checkpoint_path and completed[0] % checkpoint_every == 0:
                    _write_checkpoint(out, key_to_result, checkpoint_path)
                    log(paths, f"Checkpoint written ({completed[0]}/{len(unique_keys)} done)")

    if checkpoint_path:
        _write_checkpoint(out, key_to_result, checkpoint_path)

    retry_keys = [
        k for k, res in key_to_result.items()
        if _is_transient_error(str(res.get("occ_error") or ""))
    ]
    if retry_keys:
        log(paths, f"Retrying {len(retry_keys)} transient failures after cooldown")
        time.sleep(15)
        with ThreadPoolExecutor(max_workers=max(safe_workers // 2, 2)) as ex:
            futures = {ex.submit(fetch_one, k): k for k in retry_keys}
            for fut in tqdm(
                as_completed(futures),
                total=len(futures),
                desc="Retry occurrences",
                unit="taxon",
            ):
                k, res = fut.result()
                with result_lock:
                    key_to_result[k] = res

    return _apply_occurrence_results_vectorised(out, key_to_result)


# ---------------------------------------------------------------------------
# GBIF presence labelling and output writing
# ---------------------------------------------------------------------------

def _label_series(name_series: pd.Series, gbif_idx: pd.Series) -> pd.Series:
    cleaned    = name_series.map(lambda x: _clean_species_name(_clean_text_cell(x)))
    occurrence = cleaned.map(gbif_idx)
    in_index   = cleaned.isin(gbif_idx.index)

    result = pd.Series("implausible", index=name_series.index, dtype=object)
    result[occurrence == True]               = "plausible"    # noqa: E712
    result[~in_index]                        = "implausible"
    result[cleaned.isna() | (cleaned == "")] = None
    return result


def presence_label_for_name(gbif_idx: pd.Series, raw_name: Any) -> Optional[str]:
    raw = _clean_text_cell(raw_name)
    if raw is None:
        return None
    name = _clean_species_name(raw)
    if not name:
        return None
    if name not in gbif_idx.index:
        return "implausible"
    occurs = gbif_idx.loc[name]
    return "plausible" if occurs is True else "implausible"


def filter_other_candidates_cell(gbif_idx: pd.Series, cell: Any) -> Optional[str]:
    names = _parse_other_candidates(cell)
    kept  = [n for n in names if n in gbif_idx.index and gbif_idx.loc[n] is True]
    return "; ".join(kept) if kept else None


def write_colored_presence_xlsx(df: pd.DataFrame, out_path: Path) -> None:
    df.to_excel(out_path, index=False)
    wb = load_workbook(out_path)
    ws = wb.active

    header  = [cell.value for cell in ws[1]]
    col_map = {name: (idx + 1) for idx, name in enumerate(header)}

    presence_cols = ["GBIF_presence_species", "GBIF_presence_genus", "GBIF_presence_family"]
    green_font    = Font(color=GREEN)
    red_font      = Font(color=RED)

    for r in range(2, ws.max_row + 1):
        for cname in presence_cols:
            c = col_map.get(cname)
            if not c:
                continue
            cell = ws.cell(row=r, column=c)
            if cell.value == "plausible":
                cell.font = green_font
            elif cell.value == "implausible":
                cell.font = red_font

    wb.save(out_path)


def compare_to_input_and_write(
    paths: ProjectPaths,
    input_path: Path,
    occurrence_df: pd.DataFrame,
    out_path: Path,
) -> None:
    inp = pd.read_excel(input_path, sheet_name=0)
    for c in ["s_species", "s_genus", "s_family", "other_candidates"]:
        if c not in inp.columns:
            raise ValueError(f"Input file missing required column: {c}")

    gbif = occurrence_df.copy()

    gbif["_query_name"] = gbif["scientificName"].map(
        lambda x: _clean_species_name(_clean_text_cell(x))
    )
    gbif = gbif.dropna(subset=["_query_name"]).drop_duplicates(
        subset=["_query_name"], keep="first"
    )

    def _make_rank_idx(rank: str) -> pd.Series:
        subset = gbif[
            gbif["original_rank"].str.lower().str.strip() == rank
        ].drop_duplicates(subset=["_query_name"], keep="first")
        return subset.set_index("_query_name")["occurs_in_area"]

    idx_species = _make_rank_idx("species")
    idx_genus   = _make_rank_idx("genus")
    idx_family  = _make_rank_idx("family")

    log(paths, (
        f"GBIF index sizes — "
        f"species: {len(idx_species)}, "
        f"genus: {len(idx_genus)}, "
        f"family: {len(idx_family)}"
    ))

    inp["GBIF_presence_species"] = _label_series(inp["s_species"], idx_species)
    inp["GBIF_presence_genus"]   = _label_series(inp["s_genus"],   idx_genus)
    inp["GBIF_presence_family"]  = _label_series(inp["s_family"],  idx_family)

    inp["remaining_other_candidates"] = inp["other_candidates"].map(
        lambda cell: filter_other_candidates_cell(idx_species, cell)
    )
    inp = inp.drop(columns=["other_candidates"])

    insert_cols = [
        "GBIF_presence_family",
        "GBIF_presence_genus",
        "GBIF_presence_species",
        "remaining_other_candidates",
    ]

    cols    = list(inp.columns)
    cols_wo = [c for c in cols if c not in insert_cols]

    if "sum_of_reads" in cols_wo:
        anchor   = cols_wo.index("sum_of_reads")
        new_cols = cols_wo[:anchor] + insert_cols + cols_wo[anchor:]
        log(paths, "Inserting GBIF presence columns before 'sum_of_reads'.")
    elif "quality" in cols_wo:
        anchor   = cols_wo.index("quality")
        new_cols = cols_wo[:anchor + 1] + insert_cols + cols_wo[anchor + 1:]
        log(paths, "Inserting GBIF presence columns after 'quality'.")
    else:
        new_cols = cols_wo + insert_cols
        log(paths, "Neither 'sum_of_reads' nor 'quality' found; appending GBIF columns at end.")

    inp = inp[new_cols]
    write_colored_presence_xlsx(inp, out_path)


def find_latest_results_xlsx(paths: ProjectPaths) -> Path:
    files = sorted(
        paths.analysis_results.glob("results_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not files:
        raise FileNotFoundError(f"No results_*.xlsx found in {paths.analysis_results}")
    return files[0]


def step3_gbif(
    paths: ProjectPaths,
    gbif_settings: dict[str, Any],
    analysis_results_default: Optional[Path],
) -> None:
    """
    Orchestrate Step 3 — always uses the GBIF occurrence/search API.

    Credentials (gbif_user / gbif_pwd) are optional but recommended:
        - Without credentials: workers_unauth * 4 threads, capped at 16.
        - With credentials:    workers_auth  * 4 threads, capped at 32,
          plus more lenient rate limits from the GBIF API.
    """
    log(paths, "STEP 3: GBIF OCCURRENCE")

    inp_override = gbif_settings.get("input_results_xlsx")
    input_xlsx = (
        analysis_results_default
        if _is_empty(inp_override)
        else resolve_path_setting(paths, inp_override, paths.input)
    )
    if input_xlsx is None:
        input_xlsx = find_latest_results_xlsx(paths)
    if not input_xlsx.exists():
        raise FileNotFoundError(f"GBIF input results file not found: {input_xlsx}")

    year_from = get_int(gbif_settings,   "year_from")
    lon       = get_float(gbif_settings, "lon")
    lat       = get_float(gbif_settings, "lat")
    radius_m  = get_float(gbif_settings, "radius_m")
    min_occ   = get_int(gbif_settings,   "min_occurrences", default=5)
    n_points  = get_int(gbif_settings,   "n_points",        default=60)
    rps       = get_float(gbif_settings, "rps",             default=4.0)

    # ---------------------------------------------------------------------------
    # Resolve credentials — settings first, then environment variables
    # ---------------------------------------------------------------------------
    gbif_user = get_str(gbif_settings, "gbif_user", default="")
    gbif_pwd  = get_str(gbif_settings, "gbif_pwd",  default="")

    gbif_user = gbif_user or os.environ.get("GBIF_USER", "") or None
    gbif_pwd  = gbif_pwd  or os.environ.get("GBIF_PWD",  "") or None

    authenticated = bool(gbif_user and gbif_pwd)

    # ---------------------------------------------------------------------------
    # Derive worker count based on credential presence
    # ---------------------------------------------------------------------------
    if authenticated:
        workers_setting = get_int(gbif_settings, "workers_auth",   default=8)
        safe_workers    = min(workers_setting * 4, 32)
        log(paths, (
            f"Authenticated as '{gbif_user}' — "
            f"using {safe_workers} workers (workers_auth={workers_setting})"
        ))
    else:
        workers_setting = get_int(gbif_settings, "workers_unauth", default=4)
        safe_workers    = min(workers_setting * 4, 16)
        log(paths, (
            f"No credentials — "
            f"using {safe_workers} workers (workers_unauth={workers_setting})"
        ))

    base            = paths.gbif_results / f"gbif_{RUN_TS}"
    checkpoint_path = paths.gbif_results / f"gbif_checkpoint_{RUN_TS}.xlsx"

    # Step 3a: extract taxonomy table
    taxonomy_df   = extract_taxonomy_table_gbif(paths, input_xlsx)
    taxonomy_path = Path(str(base) + "_taxonomy_table.xlsx")
    taxonomy_df.to_excel(taxonomy_path, index=False)
    log(paths, f"Wrote {taxonomy_path}")

    # Step 3b: match taxa against GBIF backbone
    matches_df = match_taxonomy_with_gbif(
        paths, taxonomy_df,
        safe_workers=safe_workers,
        gbif_user=gbif_user,
        gbif_pwd=gbif_pwd,
    )

    # Step 3c: build WKT geometry
    wkt = build_wkt_circle(lon, lat, radius_m, n_points=n_points)

    # Step 3d: count occurrences via API
    log(paths, f"Occurrence mode: API ({safe_workers} workers, target {rps} rps)")
    occ_df = count_occurrences_api(
        paths, matches_df,
        wkt_geometry=wkt,
        year_from=year_from,
        min_occurrences=min_occ,
        safe_workers=safe_workers,
        checkpoint_path=checkpoint_path,
        gbif_user=gbif_user,
        gbif_pwd=gbif_pwd,
    )

    occ_path = Path(str(base) + "_gbif_occurrence_counts.xlsx")
    occ_df.to_excel(occ_path, index=False)
    log(paths, f"Wrote {occ_path}")

    compare_path = Path(str(base) + "_results_with_gbif_presence.xlsx")
    compare_to_input_and_write(paths, input_xlsx, occ_df, compare_path)
    log(paths, f"Wrote {compare_path}")

    if checkpoint_path.exists():
        checkpoint_path.unlink(missing_ok=True)

    log(paths, "GBIF occurrence finished.")


# ---------------------------------------------------------------------------
# CLI commands
# ---------------------------------------------------------------------------

def cmd_init(project_name: str, force: bool) -> None:
    root  = Path(project_name).resolve()
    paths = ProjectPaths(root)
    ensure_structure(paths)
    paths.logfile.write_text("", encoding="utf-8")
    write_settings_template(paths, force=force)
    log(paths, f"Project initialized at: {paths.root}")
    log(paths, "Next: put files into ./input and edit settings.xlsx.")
    log(paths, f"Run: python3 pipeline.py run --project {project_name} --steps all")


def cmd_run(project_name: str, steps: str) -> None:
    root  = Path(project_name).resolve()
    paths = ProjectPaths(root)
    ensure_structure(paths)

    if not paths.settings.exists():
        raise FileNotFoundError(
            f"Missing {paths.settings}. "
            f"Run: python3 pipeline.py init {project_name}"
        )

    blast_settings    = read_sheet_kv(paths, "blast")
    analysis_settings = read_sheet_kv(paths, "analysis")
    gbif_settings     = read_sheet_kv(paths, "gbif")

    log(paths, f"Pipeline start (RUN_TS={RUN_TS}, steps={steps})")
    log_settings(paths, "blast",    blast_settings)
    log_settings(paths, "analysis", analysis_settings)
    log_settings(paths, "gbif",     gbif_settings)
    log(paths, "------------------------")

    want = (
        {"blast", "analysis", "gbif"}
        if steps == "all"
        else {s.strip() for s in steps.split(",") if s.strip()}
    )

    blast_out:    Optional[Path] = None
    results_xlsx: Optional[Path] = None

    if "blast" in want:
        blast_out = step1_blast(paths, blast_settings)

    if "analysis" in want:
        if blast_out is None:
            candidates = sorted(
                paths.blast_results.glob("blast_*.out"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
            blast_out = candidates[0] if candidates else (
                paths.blast_results / f"blast_{RUN_TS}.out"
            )
        results_xlsx = step2_analysis(paths, analysis_settings, blast_out_default=blast_out)

    if "gbif" in want:
        step3_gbif(paths, gbif_settings, analysis_results_default=results_xlsx)

    cleanup_blast_db(paths)
    log(paths, "PIPELINE FINISHED.")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    ap  = argparse.ArgumentParser(formatter_class=argparse.RawTextHelpFormatter)
    sub = ap.add_subparsers(dest="cmd", required=True)

    p_init = sub.add_parser(
        "init",
        help="Create a new project folder with structure + template settings.xlsx",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    p_init.add_argument("project_name", help="Folder name to create (e.g. run_01)")
    p_init.add_argument("--force", action="store_true", help="Overwrite settings.xlsx if it exists")

    p_run = sub.add_parser(
        "run",
        help="Run pipeline steps inside a project folder",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    p_run.add_argument("--project", required=True, help="Project folder name/path created with init")
    p_run.add_argument(
        "--steps",
        default="all",
        help=(
            "all or comma-separated: blast,analysis,gbif\n"
            "\n"
            "blast output:\n"
            "\tblast_<timestamp>.out\n"
            "\n"
            "analysis output:\n"
            "\tfiltered_blast_results_<timestamp>.csv\n"
            "\tbest_blast_hits_<timestamp>.xlsx\n"
            "\tresults_<timestamp>.xlsx\n"
            "\n"
            "gbif output:\n"
            "\tgbif_<timestamp>_taxonomy_table.xlsx\n"
            "\tgbif_<timestamp>_gbif_occurrence_counts.xlsx\n"
            "\tgbif_<timestamp>_results_with_gbif_presence.xlsx\n"
        ),
    )

    args = ap.parse_args()

    if args.cmd == "init":
        cmd_init(args.project_name, force=args.force)
    elif args.cmd == "run":
        cmd_run(args.project, steps=args.steps)
    else:
        raise SystemExit("Unknown command")


if __name__ == "__main__":
    main()